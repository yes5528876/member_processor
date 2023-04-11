from tracemalloc import start
from unicodedata import name
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os,sys
import string
import csv
import glob
from tkinter import *   # from tkinter import Tk for Python 3.x
from tkinter import filedialog
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import *
import warnings
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border,Side,Alignment
import unittest, time, re
warnings.filterwarnings("ignore")

# # # # # Create result folder
localtime=time.localtime()
output_path = time.strftime("%m月%d日%H點%M分%S秒輸出結果/", localtime)
year=time.strftime("%Y",localtime)
民國年=str(int(year)-1911)
os.makedirs(output_path)

thin=Side(border_style="thin",color='000000')
border=Border(left=thin,right=thin,bottom=thin,top=thin)
align=Alignment(horizontal='left',vertical='center',wrap_text=True)


# # # # # Ask mode
print("請選擇功能:")
print("1.以 原始檔 製作 名冊")
print("2.以 名冊 製作 禮品名單")
print("3.以 名冊 製作 會費扣繳名單")
print("4.以 名冊 製作 會員大會名單(含簽到單)")
mode=input("請輸入1~4數字後按Enter:")
while(mode!='1' and mode!='2' and mode!='3' and mode!='4'):
    print("輸入錯誤，請重新輸入1~4的數字")
    mode=input("請輸入1~4數字後按Enter:")

# # # # # Ask input file and create df
print("請選擇名冊")
root1 = Tk()
root1.withdraw()
filename = filedialog.askopenfilename()
df = pd.read_excel(filename)  
 

# # # # # 2.以名冊製作禮品名單
if(mode=='2'):

    

    # df=df.drop('性別', axis='columns')
    # df=df.drop('職稱', axis='columns')
    # df=df.drop('備注', axis='columns')
    # df=df.drop('已簽委託書', axis='columns')
    # df=df.drop('Unnamed: 10', axis='columns')
    # df=df.drop('Unnamed: 11', axis='columns')
    # df=df.drop('Unnamed: 12', axis='columns')
    df=df[['會員編號','姓名','工號','單位','辦公室','電話']]
    df.to_excel(output_path+民國年+'全院禮品名冊.xlsx', index=False)
    
    df中興=pd.DataFrame()
    df六甲=pd.DataFrame()
    df中創=pd.DataFrame()
    df光復=pd.DataFrame()
    df沙崙=pd.DataFrame()
    df南創=pd.DataFrame()
    df生醫=pd.DataFrame()
    df其他=pd.DataFrame()

    for index,row in df.iterrows():
        if(re.search("^中興",str(df['辦公室'][index]))):
            df中興=df中興.append(row)
            continue
        if(re.search("^六甲",str(df['辦公室'][index]))):
            df六甲=df六甲.append(row)
            continue
        if(re.search("^中創",str(df['辦公室'][index]))):
            df中創=df中創.append(row)
            continue
        if(re.search("^光復",str(df['辦公室'][index]))):
            df光復=df光復.append(row)
            continue
        if(re.search("^沙崙",str(df['辦公室'][index]))):
            df沙崙=df沙崙.append(row)
            continue
        if(re.search("^南創",str(df['辦公室'][index]))):
            df南創=df南創.append(row)
            continue
        if(re.search("^新竹生醫",str(df['辦公室'][index]))):
            df生醫=df生醫.append(row)
            continue
        df其他=df其他.append(row)
    df中興=df中興.sort_values(by="單位")
    df中興=df中興[['會員編號','姓名','工號','單位','辦公室','電話']]
    df六甲=df六甲[['會員編號','姓名','工號','單位','辦公室','電話']]
    df中創=df中創[['會員編號','姓名','工號','單位','辦公室','電話']]
    df光復=df光復[['會員編號','姓名','工號','單位','辦公室','電話']]
    df沙崙=df沙崙[['會員編號','姓名','工號','單位','辦公室','電話']]
    df南創=df南創[['會員編號','姓名','工號','單位','辦公室','電話']]
    df生醫=df生醫[['會員編號','姓名','工號','單位','辦公室','電話']]
    df其他=df其他[['會員編號','姓名','工號','單位','辦公室','電話']]
    df中興.to_excel(output_path+民國年+'五一禮品(中興院區)發放名冊.xlsx', index=False)
    df六甲.to_excel(output_path+民國年+'五一禮品(六甲院區)發放名冊.xlsx', index=False)
    df中創.to_excel(output_path+民國年+'五一禮品(中創院區)發放名冊.xlsx', index=False)
    df光復.to_excel(output_path+民國年+'五一禮品(光復院區)發放名冊.xlsx', index=False)
    df沙崙.to_excel(output_path+民國年+'五一禮品(沙崙院區)發放名冊.xlsx', index=False)
    df南創.to_excel(output_path+民國年+'五一禮品(南創院區)發放名冊.xlsx', index=False)
    df生醫.to_excel(output_path+民國年+'五一禮品(新竹生醫院區)發放名冊.xlsx', index=False)
    df其他.to_excel(output_path+民國年+'五一禮品(其他地區)發放名冊.xlsx', index=False)

    #Sort 中興 unit to separate excel file
    df = pd.read_excel(output_path+民國年+'五一禮品(中興院區)發放名冊.xlsx')  
    #Sort the units into separate excel file
    now_unit=df['單位'][0]
    now_count=0
    df_temp=pd.DataFrame()
    df_minor_unit=pd.DataFrame()
    minor_unit_threshold=15
    for index,row in df.iterrows():
        if(now_unit==df['單位'][index]):
            now_count=now_count+1
            df_temp=df_temp.append(row)
        else:
            if(now_count<minor_unit_threshold):
                df_minor_unit=df_minor_unit.append(df_temp)
                now_unit=df['單位'][index]
                now_count=0
                df_temp=pd.DataFrame()
                df_temp=df_temp.append(row)
            else:
                df_temp=df_temp[['會員編號','姓名','工號','單位','辦公室','電話']]
                df_temp.to_excel(output_path+民國年+'禮品(中興院區'+df['單位'][index-1]+')發放名冊.xlsx', index=False)
                now_unit=df['單位'][index]
                now_count=0
                df_temp=pd.DataFrame()
                df_temp=df_temp.append(row)
    df_temp=df_temp[['會員編號','姓名','工號','單位','辦公室','電話']]
    df_temp.to_excel(output_path+民國年+'禮品(中興院區'+df['單位'][index-1]+')發放名冊.xlsx', index=False)
    df_minor_unit=df_minor_unit[['會員編號','姓名','工號','單位','辦公室','電話']]
    df_minor_unit.to_excel(output_path+民國年+'禮品(中興院區其他單位)發放名冊.xlsx', index=False)

    #Set the width of "Office"&"Telephone" column
    for root,ds,fs in os.walk(output_path):
        for f in fs:
            wb=load_workbook(os.path.join(root,f))
            ws=wb.active
            ws.print_title_rows = '1:1' # the first row
            流水號=0
            for row in ws.iter_rows(max_col=1):
                for cell in row:
                    cell.value=流水號
                    流水號=流水號+1
            ws.column_dimensions['B'].width=10.0    #name
            ws.column_dimensions['D'].width=10.0    #unit
            ws.column_dimensions['E'].width=25.0    #office
            ws.column_dimensions['F'].width=15.0    #telephone
            # thin=Side(border_style="thin",color='000000')
            # border=Border(left=thin,right=thin,bottom=thin)
            # align=Alignment(horizontal='left',vertical='center',wrap_text=True)
            ws['A1']="編號"
            if(re.search('全院禮品名冊',f)):
                max_c=6
            else:
                ws['G1']="簽名欄"
                max_c=7
            for row in ws.iter_rows(max_col=max_c):
                for cell in row:
                    cell.border=border
                    cell.alignment=align
            wb.save(os.path.join(root,f))

    print("名單製造完成，請至 "+output_path+" 資料夾查詢")




# # # # # 3.以名冊製作會費扣繳名單
if(mode == '3'):
    扣繳金額='600'
    # df=df.drop('辦公室', axis='columns')
    # df=df.drop('職稱', axis='columns')
    # df=df.drop('電話', axis='columns')
    # df=df.drop('備注', axis='columns')
    # df=df.drop('已簽委託書', axis='columns')
    # df=df.drop('Unnamed: 10', axis='columns')
    # df=df.drop('Unnamed: 11', axis='columns')
    # df=df.drop('Unnamed: 12', axis='columns')
    df=df[['會員編號','姓名','工號','性別','單位']]
    df=df.sort_values(by="單位")
    df.to_excel(output_path+民國年+'扣繳名冊to人力.xlsx', index=False)

    #Add border to all excel file and add "Money"&"Appendex" column
    for root,ds,fs in os.walk(output_path):
        for f in fs:
            流水號=0
            wb=load_workbook(os.path.join(root,f))
            ws=wb.active
            ws.print_title_rows = '1:1' # the first row
            # thin=Side(border_style="thin",color='000000')
            # border=Border(left=thin,right=thin,bottom=thin)
            # align=Alignment(horizontal='left',vertical='center')
            for row in ws.iter_rows(max_col=1):
                for cell in row:
                    cell.value=流水號
                    流水號=流水號+1
            for row in ws.iter_rows(min_col=7,max_col=7):
                for cell in row:
                    cell.value=扣繳金額
            ws['A1']="編號"
            ws['G1']="扣繳金額"
            ws['F1']="備註"
            for row in ws.iter_rows(max_col=7):
                for cell in row:
                    cell.border=border
                    cell.alignment=align
            wb.save(os.path.join(root,f))
    print("名單製造完成，請至 "+output_path+" 資料夾查詢")

# # # # # 4.以名冊製作會員大會名單
if(mode == '4'):
    # df=df.drop('辦公室', axis='columns')
    # df=df.drop('職稱', axis='columns')
    # df=df.drop('電話', axis='columns')
    # df=df.drop('備注', axis='columns')
    # df=df.drop('已簽委託書', axis='columns')
    # df=df.drop('Unnamed: 10', axis='columns')
    # df=df.drop('Unnamed: 11', axis='columns')
    # df=df.drop('Unnamed: 12', axis='columns')
    df=df[['會員編號','姓名','工號','性別','單位']]
    df=df.sort_values(by="單位")
    df.to_excel(output_path+民國年+'大會手冊名冊.xlsx', index=False)

    minor_unit_threshold=15

    now_unit=df['單位'][0]
    now_count=0
    df_temp=pd.DataFrame()
    df_minor_unit=pd.DataFrame()

    df = pd.read_excel(output_path+民國年+'大會手冊名冊.xlsx')  
    #Sort the units into separate excel file
    for index,row in df.iterrows():
        if(now_unit==df['單位'][index]):
            now_count=now_count+1
            df_temp=df_temp.append(row)
        else:
            if(now_count<minor_unit_threshold):
                df_minor_unit=df_minor_unit.append(df_temp)
                now_unit=df['單位'][index]
                now_count=0
                df_temp=pd.DataFrame()
                df_temp=df_temp.append(row)
            else:
                df_temp=df_temp[['會員編號','姓名','工號','性別','單位']]
                df_temp.to_excel(output_path+民國年+df['單位'][index-1]+'大會簽到表.xlsx', index=False)
                now_unit=df['單位'][index]
                now_count=0
                df_temp=pd.DataFrame()
                df_temp=df_temp.append(row)
    df_temp=df_temp[['會員編號','姓名','工號','單位','辦公室','電話']]
    df_temp.to_excel(output_path+民國年+df['單位'][index-1]+'大會簽到表.xlsx', index=False)
    df_minor_unit=df_minor_unit[['會員編號','姓名','工號','性別','單位']]
    df_minor_unit.to_excel(output_path+民國年+'其他單位大會簽到表.xlsx', index=False)
    
    #Add border to all excel file and add "Sign" column
    for root,ds,fs in os.walk(output_path):
        for f in fs:
            流水號=0
            wb=load_workbook(os.path.join(root,f))
            ws=wb.active
            ws.print_title_rows = '1:1' # the first row
            # thin=Side(border_style="thin",color='000000')
            # border=Border(left=thin,right=thin,bottom=thin,top=thin)
            # align=Alignment(horizontal='left',vertical='center')
            for row in ws.iter_rows(max_col=1):
                for cell in row:
                    cell.value=流水號
                    流水號=流水號+1
            ws['A1']="編號"
            if(re.search('大會手冊名冊',f)):
                max_c=5
            else:
                ws['F1']="簽名欄"
                max_c=6
            for row in ws.iter_rows(max_col=max_c):
                for cell in row:
                    cell.border=border
                    cell.alignment=align

            wb.save(os.path.join(root,f))

    print("名單製造完成，請至 "+output_path+" 資料夾查詢")
